from openpyxl import Workbook
import openpyxl
fname = "SGD_momentum_last.xlsx"

#book = Workbook()
#book.save(os.path.join(os.getcwd(), fname))
#"""
book = openpyxl.load_workbook(os.path.join(os.getcwd(), fname))
sheet_name = 'alexnet'
book.create_sheet(sheet_name)
sheet = book[sheet_name]

heads = ["Epoch", "Training loss", "Training accuracy", "Validation loss", "Validation accuracy","Epoch completed(sec)", "Pic number", "True labels", "Predictions", "Folder name","Folder idx"]

for i in range(len(heads)): sheet.cell(1, i+1).value =  heads[i]
#"""
